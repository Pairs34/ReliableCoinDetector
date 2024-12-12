import requests
from tabulate import tabulate
from tqdm import tqdm
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

top_count = 50  # Kaç coin alacağınızı belirleyin


def get_reliable_coins(market_cap_min=1000000000, volume_min=50000000):
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 250,
        "page": 1,
        "sparkline": "false"
    }

    response = requests.get(url, params=params)
    coins = response.json()

    reliable = []
    for coin in coins:
        if (coin["market_cap"] is not None and coin["market_cap"] >= market_cap_min) and \
                (coin["total_volume"] is not None and coin["total_volume"] >= volume_min):
            reliable.append({
                "id": coin["id"],
                "name": coin["name"],
                "symbol": coin["symbol"].upper(),
                "price": coin["current_price"],
                "market_cap": coin["market_cap"],
                "volume_24h": coin["total_volume"]
            })
    return reliable


def get_historical_data_cryptocompare(fsym="BTC", tsym="USD", limit=30):
    url = "https://min-api.cryptocompare.com/data/v2/histoday"
    params = {
        "fsym": fsym,
        "tsym": tsym,
        "limit": limit
    }
    response = requests.get(url, params=params)
    data = response.json()
    if data.get("Response") == "Success":
        return data["Data"]["Data"]
    else:
        return None


def get_2y_change(fsym="BTC", tsym="USD"):
    data = get_historical_data_cryptocompare(fsym, tsym, 730)
    if data and len(data) > 1:
        first_price = data[0]["close"]
        last_price = data[-1]["close"]
        if first_price > 0:
            change = ((last_price - first_price) / first_price) * 100
        else:
            change = 0
        return change
    return 0


def get_1m_buy_sell_ratio(fsym="BTC", tsym="USD"):
    data = get_historical_data_cryptocompare(fsym, tsym, 30)
    if data and len(data) > 0:
        buy_days = 0
        sell_days = 0
        for day in data:
            o = day["open"]
            c = day["close"]
            if c > o:
                buy_days += 1
            elif c < o:
                sell_days += 1
            else:
                buy_days += 0.5
                sell_days += 0.5

        total_days = buy_days + sell_days
        if total_days > 0:
            buy_ratio = (buy_days / total_days) * 100
            sell_ratio = (sell_days / total_days) * 100
            return buy_ratio, sell_ratio
        else:
            return 50, 50
    return 50, 50


def get_6_months_data(fsym="BTC", tsym="USD"):
    # Son 180 gün
    data = get_historical_data_cryptocompare(fsym, tsym, 180)
    if not data or len(data) == 0:
        return []

    monthly_data = {}
    for day in data:
        timestamp = day["time"]
        dt = datetime.utcfromtimestamp(timestamp)
        month_key = dt.strftime("%Y-%m")
        if month_key not in monthly_data:
            monthly_data[month_key] = []
        monthly_data[month_key].append(day)

    extended_results = []
    for month_key, records in monthly_data.items():
        buy_days = 0
        sell_days = 0
        for d in records:
            o = d["open"]
            c = d["close"]
            if c > o:
                buy_days += 1
            elif c < o:
                sell_days += 1
            else:
                buy_days += 0.5
                sell_days += 0.5
        total_days = buy_days + sell_days
        if total_days > 0:
            buy_ratio = (buy_days / total_days) * 100
            sell_ratio = (sell_days / total_days) * 100
        else:
            buy_ratio, sell_ratio = 50, 50

        dt_example = datetime.strptime(month_key, "%Y-%m")
        month_name = dt_example.strftime("%B")
        extended_results.append((dt_example, month_name, buy_ratio, sell_ratio))

    # Tarihe göre sırala
    extended_results.sort(key=lambda x: x[0])
    # Son 6 aya ihtiyacımız var
    last_6 = extended_results[-6:] if len(extended_results) > 6 else extended_results

    return last_6  # format: [(datetime, month_name, buy_ratio, sell_ratio), ...]


if __name__ == "__main__":
    pbar = tqdm(total=0, desc="Overall progress", unit="step")

    # 1. Güvenilir coinleri çek
    reliable_coins = get_reliable_coins()
    pbar.update(1)

    # top_count kadar coin al
    reliable_coins = reliable_coins[:top_count]

    # 2. Ucuz coinleri filtrele
    cheap_coins = [c for c in reliable_coins if c['price'] < 10.0]
    pbar.update(1)

    headers = [
        "Name",
        "Symbol",
        "Price($)",
        "Market Cap($)",
        "24h Volume($)",
        "Potential(%)",
        "Popularity(%)",
        "1 Month Buy/Sell Ratio",
        "2 Year Change(%)",
        "M1 Ratio",
        "M2 Ratio",
        "M3 Ratio",
        "M4 Ratio",
        "M5 Ratio",
        "M6 Ratio",
        "Trend"
    ]

    total_steps = 2 + len(cheap_coins) + 1 + 1
    pbar.total = total_steps
    pbar.refresh()

    results = []
    if not cheap_coins:
        # Tablo boş
        print(tabulate([], headers=headers, tablefmt="fancy_grid"))
        pbar.update(1)  # tablo adımı
        pbar.update(1)  # excel adımı
    else:
        avg_volume = sum(c['volume_24h'] for c in cheap_coins) / len(cheap_coins) if len(cheap_coins) > 0 else 1

        for coin in cheap_coins:
            potential = (coin['volume_24h'] / coin['market_cap']) * 100 if coin['market_cap'] != 0 else 0
            popularity = (coin['volume_24h'] / avg_volume) * 100 if avg_volume > 0 else 0
            change_2y = get_2y_change(coin['symbol'], "USD")
            buy_ratio_1m, sell_ratio_1m = get_1m_buy_sell_ratio(coin['symbol'], "USD")
            last_6 = get_6_months_data(coin['symbol'], "USD")

            buy_sell_str = f"%{round(buy_ratio_1m, 2)} buy / %{round(sell_ratio_1m, 2)} sell"

            # Ay verilerini yaz
            month_ratios = []
            count_buy_higher = 0
            for m in last_6:
                _, mname, br, sr = m
                month_ratios.append(f"{mname[:3]}: %{round(br, 2)} buy / %{round(sr, 2)} sell")
                if br > sr:
                    count_buy_higher += 1

            # Eksik aylar için boş string ekle (en eski solda)
            while len(month_ratios) < 6:
                month_ratios.insert(0, "")

            # Trend Hesaplama:
            # Son 6 ayın verisi tam 6 ay ise, en az 4 ay buy>sell ise uptrend
            # Eğer 6 aydan az veri varsa uptrend yok.
            if len(last_6) == 6 and count_buy_higher >= 4:
                trend_str = "Uptrend"
            else:
                trend_str = ""

            results.append([
                               coin['name'],
                               coin['symbol'],
                               coin['price'],
                               coin['market_cap'],
                               coin['volume_24h'],
                               round(potential, 2),
                               round(popularity, 2),
                               buy_sell_str,
                               round(change_2y, 2)
                           ] + month_ratios + [trend_str])
            pbar.update(1)  # coin işleme adımı

        # Tablo yazdır (console)
        print(tabulate(results, headers=headers, tablefmt="fancy_grid"))
        pbar.update(1)  # tablo adımı

        # Excel'e yaz
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")

        for row_idx, row_data in enumerate(results, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        for row_idx in range(2, 2 + len(results)):
            # Potential (col 6)
            pot_val = ws.cell(row=row_idx, column=6).value
            if pot_val > 100:
                ws.cell(row=row_idx, column=6).fill = green_fill

            # Popularity (col 7)
            pop_val = ws.cell(row=row_idx, column=7).value
            if pop_val > 100:
                ws.cell(row=row_idx, column=7).fill = green_fill

            # 1 Month Buy/Sell (col 8)
            ratio_str = ws.cell(row=row_idx, column=8).value
            match = re.findall(r"(\d+(\.\d+)?)", ratio_str)
            if match and len(match) >= 2:
                buy_val = float(match[0][0])
                sell_val = float(match[1][0])
                if buy_val > 100 or sell_val > 100:
                    ws.cell(row=row_idx, column=8).fill = green_fill

            # 2 Year Change (col 9)
            change_2y_val = ws.cell(row=row_idx, column=9).value
            if change_2y_val > 100:
                ws.cell(row=row_idx, column=9).fill = green_fill

            # Trend (col 16)
            trend_val = ws.cell(row=row_idx, column=16).value
            if trend_val == "Uptrend":
                ws.cell(row=row_idx, column=16).fill = red_fill

        excel_filename = "results.xlsx"
        wb.save(excel_filename)
        pbar.update(1)  # excel adımı

    pbar.close()

    print("Data saved to results.xlsx")
    os.startfile("results.xlsx")
