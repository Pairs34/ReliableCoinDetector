import requests
from tabulate import tabulate
from tqdm import tqdm
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def get_reliable_coins(market_cap_min=1000000000, volume_min=50000000):
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 250,
        "page": 1,
        "sparkline": "false"
    }

    response = requests.get(url, params=params)
    coins = response.json()

    reliable = []
    for coin in coins:
        if (coin["market_cap"] is not None and coin["market_cap"] >= market_cap_min) and \
           (coin["total_volume"] is not None and coin["total_volume"] >= volume_min):
            reliable.append({
                "id": coin["id"],
                "name": coin["name"],
                "symbol": coin["symbol"].upper(),
                "price": coin["current_price"],
                "market_cap": coin["market_cap"],
                "volume_24h": coin["total_volume"]
            })
    return reliable

def get_historical_data_cryptocompare(fsym="BTC", tsym="USD", limit=30):
    url = "https://min-api.cryptocompare.com/data/v2/histoday"
    params = {
        "fsym": fsym,
        "tsym": tsym,
        "limit": limit
    }
    response = requests.get(url, params=params)
    data = response.json()
    if data.get("Response") == "Success":
        return data["Data"]["Data"]
    else:
        return None

def get_2y_change(fsym="BTC", tsym="USD"):
    data = get_historical_data_cryptocompare(fsym, tsym, 730)
    if data and len(data) > 1:
        first_price = data[0]["close"]
        last_price = data[-1]["close"]
        if first_price > 0:
            change = ((last_price - first_price) / first_price) * 100
        else:
            change = 0
        return change
    return 0

def get_1m_buy_sell_ratio(fsym="BTC", tsym="USD"):
    data = get_historical_data_cryptocompare(fsym, tsym, 30)
    if data and len(data) > 0:
        buy_days = 0
        sell_days = 0
        for day in data:
            o = day["open"]
            c = day["close"]
            if c > o:
                buy_days += 1
            elif c < o:
                sell_days += 1
            else:
                buy_days += 0.5
                sell_days += 0.5

        total_days = buy_days + sell_days
        if total_days > 0:
            buy_ratio = (buy_days / total_days) * 100
            sell_ratio = (sell_days / total_days) * 100
            return buy_ratio, sell_ratio
        else:
            return 50, 50
    return 50, 50

if __name__ == "__main__":
    pbar = tqdm(total=0, desc="Overall progress", unit="step")

    # 1. Güvenilir coinleri çek
    reliable_coins = get_reliable_coins()
    pbar.update(1)

    # 2. Ucuz coinleri filtrele
    cheap_coins = [c for c in reliable_coins if c['price'] < 10.0]
    pbar.update(1)

    headers = [
        "Name",
        "Symbol",
        "Price($)",
        "Market Cap($)",
        "24h Volume($)",
        "Potential(%)",
        "Popularity(%)",
        "1 Month Buy/Sell Ratio",
        "2 Year Change(%)"
    ]

    total_steps = 2 + len(cheap_coins) + 1 + 1
    pbar.total = total_steps
    pbar.refresh()

    if not cheap_coins:
        # Tablo boş
        print(tabulate([], headers=headers, tablefmt="fancy_grid"))
        pbar.update(1)  # tablo adımı
        pbar.update(1)  # excel adımı
    else:
        avg_volume = sum(c['volume_24h'] for c in cheap_coins) / len(cheap_coins) if len(cheap_coins) > 0 else 1
        results = []

        for coin in cheap_coins:
            potential = (coin['volume_24h'] / coin['market_cap']) * 100 if coin['market_cap'] != 0 else 0
            popularity = (coin['volume_24h'] / avg_volume) * 100 if avg_volume > 0 else 0
            change_2y = get_2y_change(coin['symbol'], "USD")
            buy_ratio, sell_ratio = get_1m_buy_sell_ratio(coin['symbol'], "USD")

            # Bu sefer değerleri string değil, doğrudan sayısal veya sade formatta saklayalım.
            # Excel için sayısal değerler lazım, buy/sell ratio da iki ayrı değer:
            buy_sell_str = f"{round(buy_ratio,2)} buy / {round(sell_ratio,2)} sell"

            results.append([
                coin['name'],
                coin['symbol'],
                coin['price'],
                coin['market_cap'],
                coin['volume_24h'],
                round(potential,2),
                round(popularity,2),
                buy_sell_str,
                round(change_2y,2)
            ])
            pbar.update(1)  # coin işleme adımı

        # Tablo yazdır (console)
        print(tabulate(results, headers=headers, tablefmt="fancy_grid"))
        pbar.update(1) # tablo adımı

        # Excel'e yaz
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # Başlıklar
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        # Hücreleri doldur
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for row_idx, row_data in enumerate(results, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)

        # Renklendirme: %100 üzerindekilere yeşil
        # Potential(%): col 6
        # Popularity(%): col 7
        # 1 Month Buy/Sell Ratio: col 8 (format: "XX.xx buy / YY.yy sell")
        # 2 Year Change(%): col 9

        for row_idx in range(2, 2 + len(results)):
            # Potential
            pot_val = ws.cell(row=row_idx, column=6).value
            if pot_val > 100:
                ws.cell(row=row_idx, column=6).fill = green_fill

            # Popularity
            pop_val = ws.cell(row=row_idx, column=7).value
            if pop_val > 100:
                ws.cell(row=row_idx, column=7).fill = green_fill

            # 1 Month Buy/Sell Ratio
            ratio_str = ws.cell(row=row_idx, column=8).value
            # parse "XX.xx buy / YY.yy sell"
            match = re.findall(r"(\d+(\.\d+)?)", ratio_str)
            # match dönerse [("XX.xx","XX"),("YY.yy","YY")] gibi
            if match and len(match) >= 2:
                buy_val = float(match[0][0])
                sell_val = float(match[1][0])
                if buy_val > 100:
                    ws.cell(row=row_idx, column=8).fill = green_fill
                if sell_val > 100:
                    ws.cell(row=row_idx, column=8).fill = green_fill

            # 2 Year Change(%)
            change_2y_val = ws.cell(row=row_idx, column=9).value
            if change_2y_val > 100:
                ws.cell(row=row_idx, column=9).fill = green_fill

        excel_filename = "results.xlsx"
        wb.save(excel_filename)
        pbar.update(1) # excel adımı

    pbar.close()

    print("Data saved to results.xlsx")
    # Excel dosyasını aç
    os.startfile("results.xlsx")
